from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.chart import (
    BarChart,
    Reference,
    Series
)
from openpyxl.styles import (
    Font,
    Color,
    colors
)
from openpyxl.chart.layout import (
    Layout,
    ManualLayout
)


# Get time for file name
now = datetime.now()


def write(data):
    try:
        with open('Report/ViolationReport.xlsx'):
            wb = load_workbook("Report/ViolationReport.xlsx")
            ws = wb.active
    except IOError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Violation Data"
        ws["A1"] = "Violation"
        ws["B1"] = "Location"
        ws["C1"] = "Timestamp"
        ws["D1"] = "Position"
        ws["E1"] = "Occurrence"

    found = False

    for i in range(ws.min_row + 1, ws.max_row + 1):
        if data[0][1] == ws.cell(i, 1).value:
            if data[1][1] == ws.cell(i, 2).value:
                ws.cell(column=6, row=i).value += 1
                ws.cell(column=7, row=i).value += int(data[4][1])
                ws.cell(column=5, row=i).value = ws.cell(column=7, row=i).value / ws.cell(column=6, row=i).value
                ws.cell(column=5, row=i).value = round(ws.cell(column=5, row=i).value, 2)
                found = True
                break

    if not found:
        m = ws.max_row + 1
        for i in range(len(data)):
            if data[i][1].isnumeric():
                ws.cell(column=i + 1, row=m).value = float(data[i][1])
            else:
                ws.cell(column=i + 1, row=m).value = data[i][1]

        ws.cell(column=6, row=m).value = 1
        ws.cell(column=6, row=m).font = Font(color=colors.WHITE)
        ws.cell(column=7, row=m).value = float(data[4][1])
        ws.cell(column=7, row=m).font = Font(color=colors.WHITE)

    createChart(ws)

    wb.save("Report/ViolationReport.xlsx")
    print("Report compiled")


def createChart(ws):
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Violation occurrence per area"
    chart1.y_axis.title = 'Violation occurrence'
    chart1.x_axis.title = 'Violation and Location'

    cats = Reference(ws, min_col=1, max_col=2, min_row=2, max_row=ws.max_row)
    data = Reference(ws, min_col=5, max_col=5, min_row=1, max_row=ws.max_row)
    chart1.add_data(data, titles_from_data=True)
    chart1.legend = None
    chart1.set_categories(cats)
    chart1.width = 30
    chart1.height = 15
    ws.add_chart(chart1, "K2")
